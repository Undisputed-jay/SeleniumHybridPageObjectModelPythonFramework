import openpyxl

# Get total rows
def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

# Get total columns
def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column

# Get the cell data
def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row = row_number, column = column_number).value

# Write back to the cell data
def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row = row_number, column = column_number).value = data
    workbook.save(path)


def get_data_from_excel(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    
    for a in range(2, total_rows+1):
        row_list = []
        for b in range(1, total_columns+1):
            row_list.append(sheet.cell(row = a, column = b).value)
        final_list.append(row_list)
    return final_list